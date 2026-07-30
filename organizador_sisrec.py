#!/usr/bin/env python3
"""Crea estructura por FOLIO y clasifica archivos en CE o T."""

from __future__ import annotations

import argparse
import re
import shutil
import sys
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


SUBFOLDERS = ("CE", "T")


def folio_to_text(value: object) -> str | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def get_folios(excel_file: Path, sheet_name: str | None) -> list[str]:
    workbook = load_workbook(excel_file, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
    except KeyError as error:
        available = ", ".join(workbook.sheetnames)
        raise ValueError(f"Hoja no encontrada. Disponibles: {available}") from error

    header_cells = next(worksheet.iter_rows(max_row=1, values_only=True), None)
    if not header_cells:
        raise ValueError("La hoja no tiene encabezados.")
    try:
        folio_index = [str(value).strip().upper() if value else "" for value in header_cells].index("FOLIO")
    except ValueError as error:
        raise ValueError("No se encontró la columna 'FOLIO'.") from error

    folios: list[str] = []
    seen: set[str] = set()
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        folio = folio_to_text(row[folio_index] if folio_index < len(row) else None)
        if folio and folio not in seen:
            seen.add(folio)
            folios.append(folio)
    return folios


def get_coeg_targets(excel_file: Path, sheet_name: str | None) -> dict[str, str]:
    """Relaciona cada COEG_NUMERO con el primer FOLIO donde aparece."""
    workbook = load_workbook(excel_file, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
    except KeyError as error:
        available = ", ".join(workbook.sheetnames)
        raise ValueError(f"Hoja no encontrada. Disponibles: {available}") from error

    header_cells = next(worksheet.iter_rows(max_row=1, values_only=True), None)
    if not header_cells:
        raise ValueError("La hoja no tiene encabezados.")
    headers = [str(value).strip().upper() if value else "" for value in header_cells]
    try:
        folio_index = headers.index("FOLIO")
        coeg_index = headers.index("COEG_NUMERO")
    except ValueError as error:
        raise ValueError("No se encontraron las columnas FOLIO y COEG_NUMERO.") from error

    targets: dict[str, str] = {}
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        folio = folio_to_text(row[folio_index] if folio_index < len(row) else None)
        coeg_number = folio_to_text(row[coeg_index] if coeg_index < len(row) else None)
        if folio and coeg_number and coeg_number not in targets:
            targets[coeg_number] = folio
    return targets


def is_becados_agreement(value: object) -> bool:
    normalized = unicodedata.normalize("NFKD", str(value or "").casefold())
    words = set(re.findall(r"[a-z]+", "".join(char for char in normalized if not unicodedata.combining(char))))
    return {"convenio", "becados"}.issubset(words)


def get_id_targets(excel_file: Path, sheet_name: str | None) -> dict[str, list[str]]:
    """Relaciona cada ID con sus FOLIO, excepto convenios de becados."""
    workbook = load_workbook(excel_file, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
    except KeyError as error:
        available = ", ".join(workbook.sheetnames)
        raise ValueError(f"Hoja no encontrada. Disponibles: {available}") from error

    header_cells = next(worksheet.iter_rows(max_row=1, values_only=True), None)
    if not header_cells:
        raise ValueError("La hoja no tiene encabezados.")
    headers = [str(value).strip().upper() if value else "" for value in header_cells]
    try:
        id_index = headers.index("ID")
        folio_index = headers.index("FOLIO")
        document_type_index = headers.index("TIPO_DOCUMENTO")
    except ValueError as error:
        raise ValueError("No se encontraron las columnas ID, FOLIO y TIPO_DOCUMENTO.") from error

    targets: dict[str, list[str]] = {}
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if is_becados_agreement(row[document_type_index] if document_type_index < len(row) else None):
            continue
        identifier = folio_to_text(row[id_index] if id_index < len(row) else None)
        folio = folio_to_text(row[folio_index] if folio_index < len(row) else None)
        if identifier and folio:
            target_folios = targets.setdefault(identifier, [])
            if folio not in target_folios:
                target_folios.append(folio)
    return targets


def make_structure(base_folder: Path, folios: list[str], apply_changes: bool) -> None:
    for folio in folios:
        for subfolder in SUBFOLDERS:
            target = base_folder / folio / subfolder
            print(f"CREAR: {target}")
            if apply_changes:
                target.mkdir(parents=True, exist_ok=True)


def find_files(source_folder: Path, recursive: bool) -> list[Path]:
    iterator = source_folder.rglob("*") if recursive else source_folder.glob("*")
    return [path for path in iterator if path.is_file()]


def matching_folios(file_name: str, folios: list[str]) -> list[str]:
    matches = []
    for folio in folios:
        if re.search(rf"(?<!\d){re.escape(folio)}(?!\d)", file_name):
            matches.append(folio)
    return matches


def avoid_collision(destination: Path) -> Path:
    if not destination.exists():
        return destination
    number = 1
    while True:
        candidate = destination.with_name(
            f"{destination.stem} ({number}){destination.suffix}"
        )
        if not candidate.exists():
            return candidate
        number += 1


def sort_files(
    base_folder: Path,
    source_folder: Path,
    folios: list[str],
    subfolder: str,
    action: str,
    recursive: bool,
    apply_changes: bool,
) -> None:
    matched = skipped = 0
    for source in find_files(source_folder, recursive):
        matches = matching_folios(source.name, folios)
        if len(matches) != 1:
            reason = "no se detectó un folio" if not matches else f"folios ambiguos: {', '.join(matches)}"
            print(f"OMITIR: {source.name} ({reason})")
            skipped += 1
            continue

        destination = avoid_collision(base_folder / matches[0] / subfolder / source.name)
        print(f"{action.upper()}: {source} -> {destination}")
        if apply_changes:
            destination.parent.mkdir(parents=True, exist_ok=True)
            if action == "mover":
                shutil.move(str(source), str(destination))
            else:
                shutil.copy2(source, destination)
        matched += 1
    print(f"Archivos clasificados: {matched}. Omitidos: {skipped}.")


def matching_coeg(file_name: str, targets: dict[str, str]) -> list[str]:
    stem = Path(file_name).stem.strip()
    if stem in targets:
        return [stem]
    return [
        coeg_number
        for coeg_number in targets
        if re.search(rf"(?<!\d){re.escape(coeg_number)}(?!\d)", stem)
    ]


def sort_coeg_files(
    base_folder: Path,
    source_folder: Path,
    targets: dict[str, str],
    action: str,
    recursive: bool,
    apply_changes: bool,
) -> None:
    matched = skipped = 0
    for source in find_files(source_folder, recursive):
        matches = matching_coeg(source.name, targets)
        if len(matches) != 1:
            reason = "no se detectó un COEG_NUMERO" if not matches else f"COEG_NUMERO ambiguos: {', '.join(matches)}"
            print(f"OMITIR: {source.name} ({reason})")
            skipped += 1
            continue

        coeg_number = matches[0]
        destination = avoid_collision(base_folder / targets[coeg_number] / "CE" / source.name)
        print(f"{action.upper()}: {source} -> {destination}")
        if apply_changes:
            destination.parent.mkdir(parents=True, exist_ok=True)
            if action == "mover":
                shutil.move(str(source), str(destination))
            else:
                shutil.copy2(source, destination)
        matched += 1
    print(f"Archivos COEG clasificados en CE: {matched}. Omitidos: {skipped}.")


def matching_id(file_name: str, targets: dict[str, list[str]]) -> list[str]:
    stem = Path(file_name).stem.strip()
    if stem in targets:
        return [stem]
    return [
        identifier
        for identifier in targets
        if re.search(rf"(?<!\d){re.escape(identifier)}(?!\d)", stem)
    ]


def sort_id_files(
    base_folder: Path,
    source_folder: Path,
    targets: dict[str, list[str]],
    action: str,
    recursive: bool,
    apply_changes: bool,
) -> None:
    matched = skipped = copies = 0
    for source in find_files(source_folder, recursive):
        matches = matching_id(source.name, targets)
        if len(matches) != 1:
            reason = "no se detectó un ID válido" if not matches else f"IDs ambiguos: {', '.join(matches)}"
            print(f"OMITIR: {source.name} ({reason})")
            skipped += 1
            continue

        identifier = matches[0]
        target_folios = targets[identifier]
        if action == "mover" and len(target_folios) > 1:
            print(f"OMITIR: {source.name} (el ID {identifier} requiere copiarse en {len(target_folios)} carpetas T)")
            skipped += 1
            continue

        for folio in target_folios:
            destination = avoid_collision(base_folder / folio / "T" / source.name)
            print(f"{action.upper()}: {source} -> {destination}")
            if apply_changes:
                destination.parent.mkdir(parents=True, exist_ok=True)
                if action == "mover":
                    shutil.move(str(source), str(destination))
                else:
                    shutil.copy2(source, destination)
            copies += 1
        matched += 1
    print(f"Archivos ID clasificados en T: {matched}. Copias creadas: {copies}. Omitidos: {skipped}.")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Organizador para el Excel SISREC por FOLIO.")
    parser.add_argument("excel", type=Path, help="Archivo Excel con la columna FOLIO")
    parser.add_argument("destino", type=Path, help="Carpeta raíz de los folios")
    parser.add_argument("--hoja", default="base SISREC", help="Hoja a utilizar")
    parser.add_argument("--ejecutar", action="store_true", help="Aplica los cambios; sin esta opción solo simula")
    commands = parser.add_subparsers(dest="command", required=True)
    commands.add_parser("crear", help="Crea una carpeta por FOLIO, con CE y T dentro")
    sort_parser = commands.add_parser("ordenar", help="Clasifica archivos según el FOLIO incluido en el nombre")
    sort_parser.add_argument("origen", type=Path, help="Carpeta con los archivos a clasificar")
    sort_parser.add_argument("--subcarpeta", choices=SUBFOLDERS, required=True, help="Destino: CE o T")
    sort_parser.add_argument("--accion", choices=("copiar", "mover"), default="copiar")
    sort_parser.add_argument("--recursivo", action="store_true", help="Incluye subcarpetas del origen")
    coeg_parser = commands.add_parser("ordenar-coeg", help="Clasifica archivos por COEG_NUMERO en la carpeta CE")
    coeg_parser.add_argument("origen", type=Path, help="Carpeta con los archivos a clasificar")
    coeg_parser.add_argument("--accion", choices=("copiar", "mover"), default="copiar")
    coeg_parser.add_argument("--recursivo", action="store_true", help="Incluye subcarpetas del origen")
    id_parser = commands.add_parser("ordenar-id", help="Clasifica respaldos por ID en la carpeta T")
    id_parser.add_argument("origen", type=Path, help="Carpeta con los respaldos a clasificar")
    id_parser.add_argument("--accion", choices=("copiar", "mover"), default="copiar")
    id_parser.add_argument("--recursivo", action="store_true", help="Incluye subcarpetas del origen")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if not args.excel.is_file():
        print(f"Error: no existe el Excel: {args.excel}", file=sys.stderr)
        return 2
    try:
        folios = get_folios(args.excel, args.hoja)
    except (OSError, ValueError) as error:
        print(f"Error al leer el Excel: {error}", file=sys.stderr)
        return 2

    if not folios:
        print("Error: no se encontraron folios.", file=sys.stderr)
        return 2
    print(f"Folios detectados: {len(folios)}.")
    if args.command == "crear":
        make_structure(args.destino, folios, args.ejecutar)
    elif args.command == "ordenar":
        if not args.origen.is_dir():
            print(f"Error: no existe la carpeta de origen: {args.origen}", file=sys.stderr)
            return 2
        sort_files(args.destino, args.origen, folios, args.subcarpeta, args.accion, args.recursivo, args.ejecutar)
    elif args.command == "ordenar-coeg":
        if not args.origen.is_dir():
            print(f"Error: no existe la carpeta de origen: {args.origen}", file=sys.stderr)
            return 2
        try:
            targets = get_coeg_targets(args.excel, args.hoja)
        except (OSError, ValueError) as error:
            print(f"Error al leer el Excel: {error}", file=sys.stderr)
            return 2
        print(f"COEG_NUMERO únicos detectados: {len(targets)}.")
        sort_coeg_files(args.destino, args.origen, targets, args.accion, args.recursivo, args.ejecutar)
    else:
        if not args.origen.is_dir():
            print(f"Error: no existe la carpeta de origen: {args.origen}", file=sys.stderr)
            return 2
        try:
            targets = get_id_targets(args.excel, args.hoja)
        except (OSError, ValueError) as error:
            print(f"Error al leer el Excel: {error}", file=sys.stderr)
            return 2
        print(f"IDs válidos detectados: {len(targets)}.")
        sort_id_files(args.destino, args.origen, targets, args.accion, args.recursivo, args.ejecutar)
    print("Cambios aplicados." if args.ejecutar else "Simulación terminada; usa --ejecutar para aplicar cambios.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
